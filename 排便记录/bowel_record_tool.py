import os
import sys
import subprocess
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk

from openpyxl import Workbook, load_workbook


HEADERS = ["日期", "时间", "Bristol分型", "顺畅度", "颜色", "备注"]
BRISTOL_OPTIONS = [
    "1型(分离硬块,便秘)",
    "2型(块状香肠,偏硬)",
    "3型(表面裂纹,接近正常)",
    "4型(香肠状光滑,正常)",
    "5型(柔软小块,偏软)",
    "6型(糊状松散,偏稀)",
    "7型(水状,无固形物,腹泻)",
]
SMOOTHNESS_OPTIONS = ["顺畅", "一般", "费力"]
COLOR_OPTIONS = ["正常", "偏浅", "偏深", "带血", "其他"]


def get_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


EXCEL_PATH = os.path.join(get_base_dir(), "排便记录.xlsx")


def ensure_workbook():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "排便记录"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)
        return

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    current_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
    if current_headers != HEADERS:
        if ws.max_row == 1 and all(value is None for value in current_headers):
            for i, header in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=i, value=header)
            wb.save(EXCEL_PATH)
    wb.close()


def open_excel_file():
    ensure_workbook()
    try:
        if sys.platform.startswith("win"):
            os.startfile(EXCEL_PATH)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", EXCEL_PATH])
        else:
            subprocess.Popen(["xdg-open", EXCEL_PATH])
    except Exception as exc:
        messagebox.showerror("打开失败", f"无法打开Excel文件:\n{exc}")


class BowelRecordApp:
    def __init__(self, root):
        self.root = root
        self.root.title("排便记录")
        self.root.geometry("300x420")
        self.root.resizable(False, False)

        self.date_var = tk.StringVar()
        self.time_var = tk.StringVar()
        self.bristol_var = tk.StringVar(value=BRISTOL_OPTIONS[3])
        self.smoothness_var = tk.StringVar(value=SMOOTHNESS_OPTIONS[0])
        self.color_var = tk.StringVar(value=COLOR_OPTIONS[0])
        self.note_var = tk.StringVar()
        self.status_var = tk.StringVar()

        self.build_ui()
        self.prefill_datetime()

    def build_ui(self):
        frame = ttk.Frame(self.root, padding=14)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="日期").grid(row=0, column=0, sticky="w", pady=(0, 4))
        ttk.Entry(frame, textvariable=self.date_var, width=24).grid(
            row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10)
        )

        ttk.Label(frame, text="时间").grid(row=2, column=0, sticky="w", pady=(0, 4))
        ttk.Entry(frame, textvariable=self.time_var, width=24).grid(
            row=3, column=0, columnspan=2, sticky="ew", pady=(0, 10)
        )

        ttk.Label(frame, text="Bristol分型").grid(row=4, column=0, sticky="w", pady=(0, 4))
        bristol_box = ttk.Combobox(
            frame,
            textvariable=self.bristol_var,
            values=BRISTOL_OPTIONS,
            state="readonly",
            width=25,
        )
        bristol_box.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(0, 10))

        ttk.Label(frame, text="顺畅度").grid(row=6, column=0, sticky="w", pady=(0, 4))
        smooth_frame = ttk.Frame(frame)
        smooth_frame.grid(row=7, column=0, columnspan=2, sticky="w", pady=(0, 10))
        for option in SMOOTHNESS_OPTIONS:
            ttk.Radiobutton(
                smooth_frame,
                text=option,
                variable=self.smoothness_var,
                value=option,
            ).pack(side="left", padx=(0, 12))

        ttk.Label(frame, text="颜色").grid(row=8, column=0, sticky="w", pady=(0, 4))
        color_box = ttk.Combobox(
            frame,
            textvariable=self.color_var,
            values=COLOR_OPTIONS,
            state="readonly",
            width=11,
        )
        color_box.grid(row=9, column=0, sticky="w", pady=(0, 10))
        color_box.bind("<<ComboboxSelected>>", self.update_color_hint)

        self.color_hint = ttk.Label(frame, text="", foreground="#666666", wraplength=140)
        self.color_hint.grid(row=9, column=1, sticky="w", pady=(0, 10))

        ttk.Label(frame, text="备注").grid(row=10, column=0, sticky="w", pady=(0, 4))
        ttk.Entry(frame, textvariable=self.note_var, width=24).grid(
            row=11, column=0, columnspan=2, sticky="ew", pady=(0, 14)
        )

        button_frame = ttk.Frame(frame)
        button_frame.grid(row=12, column=0, columnspan=2, sticky="ew")
        ttk.Button(button_frame, text="保存", command=self.save_record).pack(
            side="left", fill="x", expand=True, padx=(0, 6)
        )
        ttk.Button(button_frame, text="打开Excel", command=open_excel_file).pack(
            side="left", fill="x", expand=True, padx=(6, 0)
        )

        ttk.Label(frame, textvariable=self.status_var, foreground="#2e7d32").grid(
            row=13, column=0, columnspan=2, sticky="w", pady=(12, 0)
        )

        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)

    def prefill_datetime(self):
        now = datetime.now()
        self.date_var.set(now.strftime("%Y-%m-%d"))
        self.time_var.set(now.strftime("%H:%M:%S"))

    def update_color_hint(self, _event=None):
        if self.color_var.get() == "带血":
            self.color_hint.config(text="建议留意身体状况,必要时咨询医生。")
        else:
            self.color_hint.config(text="")

    def save_record(self):
        date_text = self.date_var.get().strip()
        time_text = self.time_var.get().strip()
        bristol = self.bristol_var.get().strip()
        smoothness = self.smoothness_var.get().strip()
        color = self.color_var.get().strip()
        note = self.note_var.get().strip()

        if not date_text or not time_text:
            messagebox.showwarning("缺少信息", "请填写日期和时间。")
            return

        try:
            ensure_workbook()
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            ws.append([date_text, time_text, bristol, smoothness, color, note])
            wb.save(EXCEL_PATH)
            wb.close()
        except PermissionError:
            messagebox.showerror("保存失败", "Excel文件可能正在打开,请关闭后再试。")
            return
        except Exception as exc:
            messagebox.showerror("保存失败", f"无法保存记录:\n{exc}")
            return

        self.clear_form()
        self.status_var.set("已保存")
        self.root.after(2000, lambda: self.status_var.set(""))

    def clear_form(self):
        self.prefill_datetime()
        self.bristol_var.set(BRISTOL_OPTIONS[3])
        self.smoothness_var.set(SMOOTHNESS_OPTIONS[0])
        self.color_var.set(COLOR_OPTIONS[0])
        self.note_var.set("")
        self.update_color_hint()


def main():
    ensure_workbook()
    root = tk.Tk()
    BowelRecordApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
